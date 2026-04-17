"""Shared Excel export utilities used by blueprint modules."""
from io import BytesIO
from flask import Response


def xl_response(wb, filename: str) -> Response:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


def xl_header_fill():
    from openpyxl.styles import PatternFill
    return PatternFill('solid', fgColor='0F1C3A')


def xl_alt_fill():
    from openpyxl.styles import PatternFill
    return PatternFill('solid', fgColor='F0F2F8')


def xl_thin_border():
    from openpyxl.styles import Border, Side
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)


def xl_write_header(ws, headers, col_widths=None):
    from openpyxl.styles import Font, Alignment
    hf = xl_header_fill()
    thin = xl_thin_border()
    center = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=10)
        cell.fill = hf
        cell.border = thin
        cell.alignment = center
        if col_widths and ci - 1 < len(col_widths):
            ws.column_dimensions[cell.column_letter].width = col_widths[ci - 1]
    ws.freeze_panes = 'A2'


def xl_write_rows(ws, data_rows, alt=True):
    from openpyxl.styles import Alignment
    af = xl_alt_fill()
    thin = xl_thin_border()
    vcenter = Alignment(vertical='center', wrap_text=True)
    for ri, row in enumerate(data_rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.alignment = vcenter
            if alt and ri % 2 == 0:
                cell.fill = af
