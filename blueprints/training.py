from datetime import date, datetime as _dt
from flask import Blueprint, request, jsonify

from auth import login_required
from db import get_db
from export_helpers import xl_response, xl_write_header, xl_write_rows

training_bp = Blueprint('training', __name__)

TRAINING_CATEGORIES = {
    'food_safety': '食品安全',
    'fire_safety': '消防安全',
    'first_aid':   '急救訓練',
    'hygiene':     '衛生管理',
    'service':     '服務禮儀',
    'equipment':   '設備操作',
    'general':     '一般訓練',
    'other':       '其他',
}


def _init_training_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS training_records (
                    id              SERIAL PRIMARY KEY,
                    staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    course_name     TEXT NOT NULL,
                    category        TEXT NOT NULL DEFAULT 'general',
                    completed_date  DATE,
                    expiry_date     DATE,
                    certificate_no  TEXT DEFAULT '',
                    note            TEXT DEFAULT '',
                    created_at      TIMESTAMPTZ DEFAULT NOW(),
                    updated_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
    except Exception as e:
        print(f"[training_init] {e}")


@training_bp.route('/api/training/records', methods=['GET'])
@login_required
def api_training_list():
    staff_id = request.args.get('staff_id')
    category = request.args.get('category', '')
    expiring = request.args.get('expiring')
    expired  = request.args.get('expired')

    sql = """
        SELECT tr.*, ps.name AS staff_name, ps.department
        FROM training_records tr
        JOIN punch_staff ps ON tr.staff_id = ps.id
        WHERE 1=1
    """
    params = []
    if staff_id:
        sql += " AND tr.staff_id = %s"; params.append(int(staff_id))
    if category:
        sql += " AND tr.category = %s"; params.append(category)
    if expiring:
        days = int(expiring)
        sql += " AND tr.expiry_date IS NOT NULL AND tr.expiry_date <= CURRENT_DATE + INTERVAL '%s days' AND tr.expiry_date >= CURRENT_DATE"
        params.append(days)
    if expired == '1':
        sql += " AND tr.expiry_date IS NOT NULL AND tr.expiry_date < CURRENT_DATE"
    sql += " ORDER BY tr.expiry_date ASC NULLS LAST, tr.completed_date DESC"

    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()

    result = []
    today = date.today()
    for r in rows:
        d = dict(r)
        for k in ('completed_date', 'expiry_date', 'created_at', 'updated_at'):
            if d.get(k):
                d[k] = str(d[k])
        if d.get('expiry_date'):
            ed = _dt.strptime(d['expiry_date'], '%Y-%m-%d').date()
            days_left = (ed - today).days
            d['days_left'] = days_left
            d['status'] = 'expired' if days_left < 0 else 'expiring_soon' if days_left <= 60 else 'valid'
        else:
            d['days_left'] = None
            d['status'] = 'no_expiry'
        result.append(d)
    return jsonify(result)


@training_bp.route('/api/training/records', methods=['POST'])
@login_required
def api_training_create():
    b = request.get_json(force=True) or {}
    staff_id       = b.get('staff_id')
    course_name    = (b.get('course_name') or '').strip()
    category       = b.get('category', 'general')
    completed_date = b.get('completed_date') or None
    expiry_date    = b.get('expiry_date') or None
    certificate_no = (b.get('certificate_no') or '').strip()
    note           = (b.get('note') or '').strip()
    if not staff_id or not course_name:
        return jsonify({'error': '缺少必填欄位'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO training_records
              (staff_id, course_name, category, completed_date, expiry_date, certificate_no, note)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (staff_id, course_name, category, completed_date, expiry_date, certificate_no, note)).fetchone()
    return jsonify({'ok': True, 'id': row['id']})


@training_bp.route('/api/training/records/<int:rid>', methods=['PUT'])
@login_required
def api_training_update(rid):
    b = request.get_json(force=True) or {}
    with get_db() as conn:
        conn.execute("""
            UPDATE training_records SET
              course_name=%s, category=%s, completed_date=%s, expiry_date=%s,
              certificate_no=%s, note=%s, updated_at=NOW()
            WHERE id=%s
        """, (
            b.get('course_name'), b.get('category', 'general'),
            b.get('completed_date') or None, b.get('expiry_date') or None,
            b.get('certificate_no', ''), b.get('note', ''), rid
        ))
    return jsonify({'ok': True})


@training_bp.route('/api/training/records/<int:rid>', methods=['DELETE'])
@login_required
def api_training_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM training_records WHERE id=%s", (rid,))
    return jsonify({'ok': True})


@training_bp.route('/api/training/summary', methods=['GET'])
@login_required
def api_training_summary():
    """每位員工的訓練狀況摘要"""
    with get_db() as conn:
        staff_all = conn.execute(
            "SELECT id, name, department FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        records = conn.execute("""
            SELECT staff_id, category, expiry_date,
                   CASE
                     WHEN expiry_date IS NULL THEN 'no_expiry'
                     WHEN expiry_date < CURRENT_DATE THEN 'expired'
                     WHEN expiry_date <= CURRENT_DATE + INTERVAL '60 days' THEN 'expiring_soon'
                     ELSE 'valid'
                   END AS status
            FROM training_records
        """).fetchall()

    from collections import defaultdict
    by_staff = defaultdict(list)
    for r in records:
        by_staff[r['staff_id']].append(dict(r))

    result = []
    for s in staff_all:
        recs = by_staff[s['id']]
        result.append({
            'id': s['id'], 'name': s['name'], 'department': s['department'],
            'total':         len(recs),
            'valid':         sum(1 for r in recs if r['status'] in ('valid', 'no_expiry')),
            'expiring_soon': sum(1 for r in recs if r['status'] == 'expiring_soon'),
            'expired':       sum(1 for r in recs if r['status'] == 'expired'),
        })
    return jsonify(result)


@training_bp.route('/api/export/training-excel', methods=['GET'])
@login_required
def api_export_training_excel():
    import openpyxl
    from openpyxl.styles import PatternFill

    staff_id = request.args.get('staff_id', '')
    category = request.args.get('category', '')
    conds, params = ['TRUE'], []
    if staff_id:
        conds.append("tr.staff_id=%s"); params.append(int(staff_id))
    if category:
        conds.append("tr.category=%s"); params.append(category)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT tr.*, ps.name AS staff_name, ps.department, ps.employee_code
            FROM training_records tr
            JOIN punch_staff ps ON ps.id=tr.staff_id
            WHERE {' AND '.join(conds)}
            ORDER BY tr.completed_date DESC
        """, params).fetchall()

    today = date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '訓練記錄'
    headers = ['員工代碼', '姓名', '部門', '課程名稱', '類別', '完成日期', '到期日期', '剩餘天數', '證書號碼', '狀態', '備註']
    xl_write_header(ws, headers, [10, 10, 10, 24, 12, 12, 12, 10, 16, 10, 24])

    exp_fill  = PatternFill('solid', fgColor='FFEBEE')
    warn_fill = PatternFill('solid', fgColor='FFF8E1')
    data = []
    for r in rows:
        remain = ''
        status = '有效'
        if r.get('expiry_date'):
            try:
                ed = date.fromisoformat(str(r['expiry_date']))
                delta = (ed - today).days
                remain = delta
                if delta < 0:   status = '已過期'
                elif delta < 30: status = '即將到期'
            except Exception:
                pass
        data.append([
            r['employee_code'] or '', r['staff_name'], r['department'] or '',
            r['course_name'] or '',
            TRAINING_CATEGORIES.get(r['category'] or 'other', r['category'] or ''),
            str(r['completed_date']) if r.get('completed_date') else '',
            str(r['expiry_date'])    if r.get('expiry_date')    else '',
            remain, r['certificate_no'] or '', status, r['note'] or '',
        ])
    xl_write_rows(ws, data)

    ws2 = wb.create_sheet('到期摘要')
    xl_write_header(ws2, ['員工', '課程', '類別', '到期日', '剩餘天數', '狀態'], [12, 24, 12, 12, 10, 10])
    exp_data = [r for r in data if r[9] in ('已過期', '即將到期')]
    xl_write_rows(ws2, [[r[1], r[3], r[4], r[6], r[7], r[9]] for r in exp_data])

    return xl_response(wb, 'training_records.xlsx')
